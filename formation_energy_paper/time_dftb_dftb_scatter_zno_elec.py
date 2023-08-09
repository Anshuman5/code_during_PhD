# Import necessary libraries
from __future__ import print_function
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np  
import matplotlib.ticker as ticker
import matplotlib.transforms as mtransforms
from openpyxl import load_workbook

# Define colors and styles for data visualization
dft_color_zno='#191970'
znorg_color='r'
znorg_label= 'znorg-0-1'

# Define font sizes for labels and ticks
ylabel_fsize = 24
xlabel_fsize = 24
ylabel_fsize_bar = 24
xlabel_fsize_bar = 24
ytick_labelsize = 20
xtick_labelsize = 20
ytick_labelsize_bar = 18
xtick_labelsize_bar = 18
legend_prop_size = 24
annotation_fontsize = 22

# Create a new Excel file
writer = pd.ExcelWriter('zno_formE_5.xlsx', engine='xlsxwriter')
writer.save()

# Load the existing Excel file and sheet
wb = load_workbook('zno_formE_5.xlsx')
ws = wb['Sheet1']

# Set column headers in the sheet
ws['A1'] = 'Struc'
ws['B1'] = 'natoms'
ws['C1'] = 'nk'
ws['D1'] = 'nZn'
ws['E1'] = 'nO'
ws['F1'] = 'comp'
ws['G1'] = 'dft_total_energy'
ws['H1'] = 'dftb_total_energy_znorg'
ws['I1'] = 'dft_formE'
ws['J1'] = 'dftb_formE'

# Function to calculate the formation energy
def calc_formE(df, x1, x2, cname, formname, col):
    comp = df.comp
    nk = df.nk
    x1 = x1
    x2 = x2
    totalE = df[cname]
    e_ref_x1 = float(totalE[2])
    e_ref_x2 = float(totalE[0])
    erefx = np.zeros((len(comp)))
    ll = []
    for i in range(len(comp)):
        erefx[i] = e_ref_x1 + (comp[i]-x1)*((e_ref_x2-e_ref_x1)/(x2-x1)) 
        ll.append(totalE[i]/nk[i] - erefx[i])
    df[formname] = ll

# Function to copy data to the Excel sheet
def copy_sheet(df, cname, cletter):
    comp = df.comp
    for i in range(len(comp)):
        ws[cletter+str(i+2)] = df[cname][i]

# Read data from an existing Excel file and sheet
sheet_name = 'Sheet2'
dfe = pd.read_excel('zno_formE_4.xlsx', sheet_name=sheet_name, usecols=range(0,16))

# Calculate formation energies using the calc_formE function
calc_formE(dfe, 0, 1, 'dft_total_energy', 'formation_energy_dft', 'I')
calc_formE(dfe, 0, 1, 'dftb_total_energy_znorg', 'formation_energy_dftb', 'J')

# Get the formation energies and composition data from the DataFrame
formE_dft = dfe['formation_energy_dft']
formE_dftb = dfe['formation_energy_dftb']
comp = dfe.comp

# Copy the data to the Excel sheet
copy_sheet(dfe, 'Struc', 'A')
copy_sheet(dfe, 'natoms', 'B')
copy_sheet(dfe, 'nk', 'C')
copy_sheet(dfe, 'nZn', 'D')
copy_sheet(dfe, 'nO', 'E')
copy_sheet(dfe, 'comp', 'F')
copy_sheet(dfe, 'dft_total_energy', 'G')
copy_sheet(dfe, 'dftb_total_energy_znorg', 'H')
copy_sheet(dfe, 'formation_energy_dft', 'I')
copy_sheet(dfe, 'formation_energy_dftb', 'J')

# Save the Excel file
wb.save('zno_formE_5.xlsx')

# Calculate the root mean square error (rms_znorg) between formation energies
rms_znorg = ((formE_dft - formE_dftb) ** 2).mean() ** .5

# Print the error in formation energy with znorg
print( 'Error in formation energy with znorg (eV)', rms_znorg)

# plot the data
plt.rc('xtick', labelsize=xtick_labelsize)
plt.rc('ytick', labelsize=ytick_labelsize)
fig, axs = plt.subplots(1, 2, figsize=(21,7))
axs[0].scatter(comp,formE_dft, color=dft_color_zno, label='DFT')
axs[1].scatter(comp,formE_dftb, color=znorg_color, label='DFTB ('+znorg_label+')')
axs[0].set(xlabel ='$x$ in Zn$_{2}$O$_{2(1-x)}$')
axs[0].xaxis.get_label().set_fontsize(xlabel_fsize)
axs[1].set(xlabel ='$x$ in Zn$_{2}$O$_{2(1-x)}$')
axs[1].xaxis.get_label().set_fontsize(xlabel_fsize)
#axs[1].yaxis.set_major_locator(ticker.MultipleLocator(30))
axs[0].set(ylabel='Formation Energy (eV/unit cell)')
axs[0].yaxis.get_label().set_fontsize(ylabel_fsize)
axs[0].legend(loc='upper right', bbox_to_anchor=(0, 1.08, 1, 0.05), fancybox=True, shadow=True, handlelength=0.4, prop={'size': legend_prop_size})
axs[1].legend(loc='upper right', bbox_to_anchor=(0, 1.08, 1, 0.05), fancybox=True, shadow=True, handlelength=0.4, prop={'size': legend_prop_size})
axs[0].set_ylim([0, .9])
axs[1].set_ylim([0, .9])
unique_comp = np.sort(comp.unique().tolist())
dft_hull = []
dftb_hull_znorg = []
for i in range (len(unique_comp)):
      dft_hull.append(dfe.loc[dfe['comp']==(unique_comp[i])]['formation_energy_dft'].min())
      dftb_hull_znorg.append(dfe.loc[dfe['comp']==(unique_comp[i])]['formation_energy_dftb'].min())
#axs[0].plot([unique_comp[i] for i in (0, 9, 10, 12)], [dft_hull[i] for i in (0, 9, 10, 12)], color=dft_color_zno, label = 'DFT hull')
#axs[1].plot([unique_comp[i] for i in (0, 9, 10, 12)], [dftb_hull_znorg[i] for i in (0, 9, 10, 12)], color=znorg_color, label = 'DFTB hull ' +znorg_label)


labels = ['(a)', '(b)']
axs = axs.flat
for n, ax in enumerate(axs):
    # label physical distance to the left and up:
    trans = mtransforms.ScaledTranslation(-20/72, 7/72, fig.dpi_scale_trans)
    if n == 0:
        ax.text(-0.1, 1.05, labels[n], transform=ax.transAxes + trans,
            size=24, weight='regular', va='bottom', fontfamily='serif')
    else:
        ax.text(-0.07, 1.05, labels[n], transform=ax.transAxes + trans,
            size=24, weight='regular', va='bottom', fontfamily='serif')

# plt.savefig("zno_hull_dft_dftb.png", dpi=400, bbox_inches="tight", pad_inches=0.25)

# Time Scattr chart
sheet_name = 'Sheet2'
df = pd.read_excel('time_dft_dftb_2.xlsx', sheet_name=sheet_name, usecols=range(0,32))
time_dft_zno =  df['total_time_dft_zno_tol']
total_time_dft_zno_tol =  sum(df['total_time_dft_zno_tol'])
time_dftb_zno =  df['total_time_dftb_zno']
total_time_dftb_zno =  sum(df['total_time_dftb_zno'])
comp = df['comp']

fig, axs = plt.subplots(1, 2, figsize=(21,7))
axs[0].scatter(comp,time_dft_zno, color=dft_color_zno, label='DFT')
axs[1].scatter(comp,time_dft_zno/time_dftb_zno, color=znorg_color, label='DFTB ('+znorg_label+')')
axs[0].set(ylabel='Time (sec)')
axs[0].yaxis.get_label().set_fontsize(ylabel_fsize)
axs[1].set(ylabel='DFT Time/DFTB Time')
axs[1].yaxis.get_label().set_fontsize(ylabel_fsize)
axs[0].set(xlabel ='$x$ in Zn$_{2}$O$_{2(1-x)}$')
axs[0].xaxis.get_label().set_fontsize(xlabel_fsize)
axs[1].set(xlabel ='$x$ in Zn$_{2}$O$_{2(1-x)}$' )
axs[1].xaxis.get_label().set_fontsize(xlabel_fsize)
axs[0].legend(loc='upper right', bbox_to_anchor=(0, 1.02, 1, 0.05), fancybox=True, handlelength=0.4, shadow=True, prop={'size': legend_prop_size})
axs[1].legend(loc='upper right', bbox_to_anchor=(0, 1.02, 1, 0.05), fancybox=True, handlelength=0.4, shadow=True,prop={'size': legend_prop_size})


# Define labels for subplots
labels = ['(a)', '(b)']

# Flatten the axs object to iterate over the subplots
axs = axs.flat

# Iterate over the subplots and add labels
for n, ax in enumerate(axs):
    # Define translation for the labels
    trans = mtransforms.ScaledTranslation(-20/72, 7/72, fig.dpi_scale_trans)
    if n == 0:
        # Add label to the first subplot
        ax.text(-0.15, 1.05, labels[n], transform=ax.transAxes + trans,
            size=24, weight='regular', va='bottom', fontfamily='serif')
    else:
        # Add label to the second subplot
        ax.text(-0.07, 1.05, labels[n], transform=ax.transAxes + trans,
            size=24, weight='regular', va='bottom', fontfamily='serif')

# Save the plot as an image
# plt.savefig("zno_time_dft_dftb.png", dpi=400, bbox_inches="tight", pad_inches=0.25)

# Print the total DFT and DFTB times for ZnO
print('total dft time for ZnO (8 cpus): {0:.2f} days'.format(total_time_dft_zno_tol/(3600*24)))
print('total dftb time for ZnO (8 cpus): {0:.2f} days'.format(total_time_dftb_zno/(3600*24)))

#ZnO dft vs dftb; # atoms
# Get the number of electrons in ZnO configurations
nelectrons_zno = df['nelec_dft']
unique_nelectrons_zno = np.sort(nelectrons_zno.unique().tolist())

# Create dictionaries to store the average time per SCF cycle for DFT and DFTB
dft_time_zno_dict= {}
dftb_time_zno_dict = {}

# Iterate over the unique number of electrons in ZnO configurations
for j in unique_nelectrons_zno:
    # Initialize variables
    timezno_dft = 0
    timezno_dftb = 0
    count = 0
    
    # Iterate over the ZnO configurations
    for i in range(len(comp)):
        if nelectrons_zno[i] == j:
            count += 1
            timezno_dft += df['total_time_dft_zno_tol'][i] / df['iter_dft'][i]
            timezno_dftb += df['total_time_dftb_zno'][i] / df['iter_dftb'][i]
    
    # Calculate the average time per SCF cycle for DFT and DFTB
    dft_time_zno_dict[j] = timezno_dft / count
    dftb_time_zno_dict[j] = timezno_dftb / count

# Set the width of the bars
width = 0.9

# Get the x-axis values
xaxis_vals_dft = list(dft_time_zno_dict.keys())
xaxis_vals_dftb = list(dftb_time_zno_dict.keys())
x = range(len(xaxis_vals_dft))


plt.rc('xtick', labelsize=xtick_labelsize_bar+4)
plt.rc('ytick', labelsize=ytick_labelsize_bar+4)

plt.figure(figsize=(18, 9))

plt.xticks(x)
plt.xlabel('No. of Electrons in ZnO Configs.', fontsize=36)

plt.ylabel('Time/SCF Cycle (sec/iter.)', fontsize=36)

# Plot the bar chart for DFT
plt.bar(x, (dft_time_zno_dict.values()), tick_label=xaxis_vals_dft, width=width, color=dft_color_zno, label='DFT')

# Plot the bar chart for DFTB
plt.bar(x, (dftb_time_zno_dict.values()), tick_label=xaxis_vals_dftb, width=width, color=znorg_color, label='DFTB ('+znorg_label+')')

plt.legend(loc='upper left', fancybox=True, shadow=True, handlelength=0.8, prop={'size': legend_prop_size})

plt.show()
