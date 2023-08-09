# Import necessary libraries
from __future__ import print_function
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np  
import matplotlib.ticker as ticker
import matplotlib.transforms as mtransforms
from openpyxl import load_workbook

# Define colors and styles for data visualization
dft_color_zno='#191970'
znorg_color='r'
znorg_label= 'znorg-0-1'

# Define font sizes for labels and ticks
ylabel_fsize = 24
xlabel_fsize = 24
ylabel_fsize_bar = 24
xlabel_fsize_bar = 24
ytick_labelsize = 20
xtick_labelsize = 20
ytick_labelsize_bar = 18
xtick_labelsize_bar = 18
legend_prop_size = 24
annotation_fontsize = 22

# Create a new Excel file
writer = pd.ExcelWriter('zno_formE_5.xlsx', engine='xlsxwriter')
writer.save()

# Load the existing Excel file and sheet
wb = load_workbook('zno_formE_5.xlsx')
ws = wb['Sheet1']

# Set column headers in the sheet
ws['A1'] = 'Struc'
ws['B1'] = 'natoms'
ws['C1'] = 'nk'
ws['D1'] = 'nZn'
ws['E1'] = 'nO'
ws['F1'] = 'comp'
ws['G1'] = 'dft_total_energy'
ws['H1'] = 'dftb_total_energy_znorg'
ws['I1'] = 'dft_formE'
ws['J1'] = 'dftb_formE'

# Function to calculate the formation energy
def calc_formE(df, x1, x2, cname, formname, col):
    comp = df.comp
    nk = df.nk
    x1 = x1
    x2 = x2
    totalE = df[cname]
    e_ref_x1 = float(totalE[2])
    e_ref_x2 = float(totalE[0])
    erefx = np.zeros((len(comp)))
    ll = []
    for i in range(len(comp)):
        erefx[i] = e_ref_x1 + (comp[i]-x1)*((e_ref_x2-e_ref_x1)/(x2-x1)) 
        ll.append(totalE[i]/nk[i] - erefx[i])
    df[formname] = ll

# Function to copy data to the Excel sheet
def copy_sheet(df, cname, cletter):
    comp = df.comp
    for i in range(len(comp)):
        ws[cletter+str(i+2)] = df[cname][i]

# Read data from an existing Excel file and sheet
sheet_name = 'Sheet2'
dfe = pd.read_excel('zno_formE_4.xlsx', sheet_name=sheet_name, usecols=range(0,16))

# Calculate formation energies using the calc_formE function
calc_formE(dfe, 0, 1, 'dft_total_energy', 'formation_energy_dft', 'I')
calc_formE(dfe, 0, 1, 'dftb_total_energy_znorg', 'formation_energy_dftb', 'J')

# Get the formation energies and composition data from the DataFrame
formE_dft = dfe['formation_energy_dft']
formE_dftb = dfe['formation_energy_dftb']
comp = dfe.comp

# Copy the data to the Excel sheet
copy_sheet(dfe, 'Struc', 'A')
copy_sheet(dfe, 'natoms', 'B')
copy_sheet(dfe, 'nk', 'C')
copy_sheet(dfe, 'nZn', 'D')
copy_sheet(dfe, 'nO', 'E')
copy_sheet(dfe, 'comp', 'F')
copy_sheet(dfe, 'dft_total_energy', 'G')
copy_sheet(dfe, 'dftb_total_energy_znorg', 'H')
copy_sheet(dfe, 'formation_energy_dft', 'I')
copy_sheet(dfe, 'formation_energy_dftb', 'J')

# Save the Excel file
wb.save('zno_formE_5.xlsx')

# Calculate the root mean square error (rms_znorg) between formation energies
rms_znorg = ((formE_dft - formE_dftb) ** 2).mean() ** .5

print( 'Error in formation energy with znorg (eV)', rms_znorg)


plt.rc('xtick', labelsize=xtick_labelsize)
plt.rc('ytick', labelsize=ytick_labelsize)
fig, axs = plt.subplots(1, 2, figsize=(21,7))
axs[0].scatter(comp,formE_dft, color=dft_color_zno, label='DFT')
axs[1].scatter(comp,formE_dftb, color=znorg_color, label='DFTB ('+znorg_label+')')
axs[0].set(xlabel ='$x$ in Zn$_{2}$O$_{2(1-x)}$')
axs[0].xaxis.get_label().set_fontsize(xlabel_fsize)
axs[1].set(xlabel ='$x$ in Zn$_{2}$O$_{2(1-x)}$')
axs[1].xaxis.get_label().set_fontsize(xlabel_fsize)
#axs[1].yaxis.set_major_locator(ticker.MultipleLocator(30))
axs[0].set(ylabel='Formation Energy (eV/unit cell)')
axs[0].yaxis.get_label().set_fontsize(ylabel_fsize)
axs[0].legend(loc='upper right', bbox_to_anchor=(0, 1.08, 1, 0.05), fancybox=True, shadow=True, handlelength=0.4, prop={'size': legend_prop_size})
axs[1].legend(loc='upper right', bbox_to_anchor=(0, 1.08, 1, 0.05), fancybox=True, shadow=True, handlelength=0.4, prop={'size': legend_prop_size})
axs[0].set_ylim([0, .9])
axs[1].set_ylim([0, .9])
unique_comp = np.sort(comp.unique().tolist())
dft_hull = []
dftb_hull_znorg = []
for i in range (len(unique_comp)):
      dft_hull.append(dfe.loc[dfe['comp']==(unique_comp[i])]['formation_energy_dft'].min())
      dftb_hull_znorg.append(dfe.loc[dfe['comp']==(unique_comp[i])]['formation_energy_dftb'].min())
#axs[0].plot([unique_comp[i] for i in (0, 9, 10, 12)], [dft_hull[i] for i in (0, 9, 10, 12)], color=dft_color_zno, label = 'DFT hull')
#axs[1].plot([unique_comp[i] for i in (0, 9, 10, 12)], [dftb_hull_znorg[i] for i in (0, 9, 10, 12)], color=znorg_color, label = 'DFTB hull ' +znorg_label)


labels = ['(a)', '(b)']
axs = axs.flat
for n, ax in enumerate(axs):
    # label physical distance to the left and up:
    trans = mtransforms.ScaledTranslation(-20/72, 7/72, fig.dpi_scale_trans)
    if n == 0:
        ax.text(-0.1, 1.05, labels[n], transform=ax.transAxes + trans,
            size=24, weight='regular', va='bottom', fontfamily='serif')
    else:
        ax.text(-0.07, 1.05, labels[n], transform=ax.transAxes + trans,
            size=24, weight='regular', va='bottom', fontfamily='serif')

plt.savefig("zno_hull_dft_dftb.png", dpi=400, bbox_inches="tight", pad_inches=0.25)

# Time Scattr chart
sheet_name = 'Sheet2'
df = pd.read_excel('time_dft_dftb_2.xlsx', sheet_name=sheet_name, usecols=range(0,32))
# Extracting the relevant columns from the dataframe
time_dft_zno = df['total_time_dft_zno_tol']
total_time_dft_zno_tol = sum(df['total_time_dft_zno_tol'])
time_dftb_zno = df['total_time_dftb_zno']
total_time_dftb_zno = sum(df['total_time_dftb_zno'])
comp = df['comp']

# Creating the figure and axes for the subplots
fig, axs = plt.subplots(1, 2, figsize=(21,7))

# Plotting the scatter plot for DFT time vs. composition
axs[0].scatter(comp, time_dft_zno, color=dft_color_zno, label='DFT')
# Plotting the scatter plot for DFT time / DFTB time vs. composition
axs[1].scatter(comp, time_dft_zno/time_dftb_zno, color=znorg_color, label='DFTB (' + znorg_label + ')')

# Setting the labels for the y-axis
axs[0].set(ylabel='Time (sec)')
axs[0].yaxis.get_label().set_fontsize(ylabel_fsize)
axs[1].set(ylabel='DFT Time/DFTB Time')
axs[1].yaxis.get_label().set_fontsize(ylabel_fsize)

# Setting the labels for the x-axis
axs[0].set(xlabel ='$x$ in Zn$_{2}$O$_{2(1-x)}$')
axs[0].xaxis.get_label().set_fontsize(xlabel_fsize)
axs[1].set(xlabel ='$x$ in Zn$_{2}$O$_{2(1-x)}$' )
axs[1].xaxis.get_label().set_fontsize(xlabel_fsize)

# Adding legends to the plots
axs[0].legend(loc='upper right', bbox_to_anchor=(0, 1.02, 1, 0.05), fancybox=True, handlelength=0.4, shadow=True, prop={'size': legend_prop_size})
axs[1].legend(loc='upper right', bbox_to_anchor=(0, 1.02, 1, 0.05), fancybox=True, handlelength=0.4, shadow=True, prop={'size': legend_prop_size})

# Adding labels to the subplots
labels = ['(a)', '(b)']
axs = axs.flat
for n, ax in enumerate(axs):
    # Defining the translation for the labels
    trans = mtransforms.ScaledTranslation(-20/72, 7/72, fig.dpi_scale_trans)

    # Adding the labels to the subplots
    if n == 0:
        ax.text(-0.15, 1.05, labels[n], transform=ax.transAxes + trans,
            size=24, weight='regular', va='bottom', fontfamily='serif')
    else:
        ax.text(-0.07, 1.05, labels[n], transform=ax.transAxes + trans,
            size=24, weight='regular', va='bottom', fontfamily='serif')

# Saving the figure
plt.savefig("zno_time_dft_dftb.png", dpi=400, bbox_inches="tight", pad_inches=0.25)

# Printing the total DFT and DFTB times for ZnO (8 cpus)
print('Total DFT time for ZnO (8 cpus): {0:.2f} days'.format(total_time_dft_zno_tol / (3600*24)))
print('Total DFTB time for ZnO (8 cpus): {0:.2f} days'.format(total_time_dftb_zno / (3600*24)))

# Calculating the average DFT and DFTB times for different number of atoms in ZnO
natoms_zno = df['natoms_zno']
unique_natoms_zno = np.sort(natoms_zno.unique().tolist())
dft_time_zno_dict = {}
dftb_time_zno_dict = {}

# Looping over the unique number of atoms in ZnO
for j in unique_natoms_zno:
    timezno_dft = 0
    timezno_dftb = 0
    count = 0
    # Looping over the data to calculate the sum of DFT and DFTB times for the current number of atoms
    for i in range(len(comp)):
        if natoms_zno[i] == j:
            count += 1
            timezno_dft += df['total_time_dft_zno_tol'][i]
            timezno_dftb += df['total_time_dftb_zno'][i]
    # Calculating the average DFT and DFTB times for the current number of atoms
    dft_time_zno_dict[j] = timezno_dft / (60 * count)
    dftb_time_zno_dict[j] = timezno_dftb / (60 * count)
plt.figure(figsize=(18, 9))
plt.title('ZnO', fontsize=24)
plt.xlabel('No. of Atoms', fontsize=24)
plt.ylabel('Time (mins)',  fontsize=28)
plt.bar(dft_time_zno_dict.keys(),(dft_time_zno_dict.values()), color = dft_color_zno, label='DFT')
plt.bar(dftb_time_zno_dict.keys(),(dftb_time_zno_dict.values()), color=znorg_color, label='DFTB ('+znorg_label+')')
plt.legend(loc='upper left', fancybox=True, shadow=True, handlelength=0.4, prop={'size': legend_prop_size})
plt.xlim((1,16))
plt.show()